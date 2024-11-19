import sys
import os
curr_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(curr_dir)

import json
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from ExcelRW.readcsv import read_csv

def artist_lookup(artist_name, artist_info):
    '''
    This lookup does a more in-depth search for an artist if they were
    not found immediately.

    @param artist_name: The name of the artist we are looking for
    @param artist_info: The json containing all artist information

    @return The key of the artist if found, None otherwise
    '''

    for key, value in artist_info.items():

        if artist_name in value["aliases"] or artist_name == value["name"]:
            return key
        
    return None

def item_lookup(item_name, item_info):
    '''
    This function looks up an item that was labeled as YNM Vended. If it exists
    then we return the artists, if not then we return None.

    @param item_name: The name of the item we are looking for
    @param item_info: The json containing all item information

    @return The artist name associated with the item
    '''
    return item_info.get(item_name, None)


def name_extract(product_name, artist_info):
    '''
    This function will work to extract the collaborator's name from our artist info dictionary

    @param product_name: The name of the product we are looking at
    @param artist_info: The json containing all artist information

    @return The name of the collaborator, if we cannot find it, we consider this an imporperly named item and return None

    '''

    #Alias search the item to collab name
    for key, value in artist_info.items():

        for val in value["aliases"]:

            if val.lower() in product_name.lower():

                return key

    return None

def main():

    ynm_file_path = '/YNM/PayoutAutomator/Data/SheetPreprocessor/YNM_Sales_Final.csv'
    ynm_financial_info = read_csv(ynm_file_path)

    yne_file_path = '/YNM/PayoutAutomator/Data/SheetPreprocessor/YNE_Sales_Final.csv'
    yne_financial_info = read_csv(yne_file_path)

    rollover_path = "/YNM/PayoutAutomator/Data/PayoutPrototype/Rollovers.csv"
    rollover_info = read_csv(rollover_path)

    pending_rollover_path = '/YNM/PayoutAutomator/Data/PayoutPrototype/Pending_Rollovers.csv'
    pending_rollover_info = read_csv(pending_rollover_path)

    carry_pending_rollovers = []

    #Determine what to do with pending rollovers
    for pending_rollover in pending_rollover_info:

        if pending_rollover[8].lower() == "y":

            if pending_rollover[0].upper() == "YNM":

                ynm_financial_info.append([pending_rollover[2], pending_rollover[1], pending_rollover[3], '', '', '', '', '', '', '', pending_rollover[4], pending_rollover[5], pending_rollover[6], pending_rollover[7]])

            elif pending_rollover[0].upper() == "YNE":

                yne_financial_info.append([pending_rollover[2], pending_rollover[1], pending_rollover[3], '', '', '', '', '', '', '', pending_rollover[4], pending_rollover[5], pending_rollover[6], pending_rollover[7]])
        
        else:
            
            carry_pending_rollovers.append(pending_rollover)

    with open('/YNM/PayoutAutomator/Data/artists.json') as fp:
        artist_info = json.load(fp)

    with open('/YNM/PayoutAutomator/Data/namedItems.json') as fp:
        item_info = json.load(fp)

    ynm_original_dict = {}
    ynm_collab_dict = {}

    yne_original_dict = {}
    yne_collab_dict = {}

    #Iterate through each product that was sold this month for YNM
    for product in ynm_financial_info:

        product_vendor  = product[1].strip("'").replace("''", "'")
        product_name = product[0].strip("'").replace("''", "'") #Clean up the string to be readable in json
        distribution_type = product[2]
        product_type = product[3]
        net_quantity = product[4]
        gross_sales = product[5]
        discounts = product[6]
        returns = product[7]
        net_sales = product[8]
        taxes = product[9]
        total_sales = float(product[10])
        processor_fee = float(product[11])
        total_cost = float(product[12])
        gross_profit = total_sales - processor_fee - total_cost

        #Conduct safer search for artist, considering the fact there might be ( )
        if "(" in product_vendor:

            #Obtain all chars up to the " (""
            product_vendor = product_vendor[:product_vendor.index("(") - 1]

        #First check to see if our vendor is an *actual* artist
        if product_vendor.title() in artist_info:
            
            product_vendor = product_vendor.title()

        else:
            

            #If our quick search of artist is not found, we need to look up
            #the artist aliases. LOOK UP EXACT MATCH (No titles)
            search_vendor = artist_lookup(product_vendor, artist_info)

            if search_vendor is not None:
                
                product_vendor = search_vendor
                

            #Our last resort search will be for Yorunmachi vended items,
            #If we cannot find the vendor at this point, we can assume we 
            #need to enter a new entry or make a fix
            else:

                search_vendor_item = item_lookup(product_name, item_info)

                if search_vendor_item:

                    product_vendor = search_vendor_item
                
                else:

                    print("YNM ARTIST NOT FOUND, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                    continue
        
        #
        #Get associated information for the artist's product - handle differently dependent on the distribution type
        #

        if distribution_type == "Original":

            if product_vendor not in ynm_original_dict:

                ynm_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]

            else:

                ynm_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])

        elif distribution_type == "Commercial":

    
            if product_vendor not in ynm_collab_dict:

                ynm_collab_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]

            else:

                ynm_collab_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])


        elif distribution_type == "Collab":


            #We find the collaborator and add them to the collab dict
            collab_name = name_extract(product_name, artist_info)

            if not(collab_name):

                print("COLLABORATOR NOT FOUND, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                continue


            if collab_name not in ynm_collab_dict:

                ynm_collab_dict[collab_name] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]
            
            else:
                    
                ynm_collab_dict[collab_name].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])

            #Handle if YNM was a vendor - we have to take collaborator's name out of list as it was already accounted for
            if type(product_vendor) == list:

                try:
                    product_vendor.remove(collab_name)
                except:

                    print("FIX NAME:", collab_name, product_vendor, product_name)

                #Recreate the one element list to a string
                product_vendor = product_vendor[0]


            #Next, we must find the product vendor as an original source, so they go in the original source dict
            if product_vendor not in ynm_original_dict:

                ynm_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]

            else:

                ynm_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])  

        #TODO - Handle Bundles
        elif distribution_type == "Bundle":

            pass   


    #Iterate through each product that was sold this month for YNE
    for product in yne_financial_info:

        product_vendor  = product[1].strip("'").replace("''", "'")
        product_name = product[0].strip("'").replace("''", "'") #Clean up the string to be readable in json
        distribution_type = product[2]
        product_type = product[3]
        net_quantity = product[4]
        gross_sales = product[5]
        discounts = product[6]
        returns = product[7]
        net_sales = product[8]
        taxes = product[9]
        total_sales = float(product[10])
        processor_fee = float(product[11])
        total_cost = float(product[12])
        gross_profit = total_sales - processor_fee - total_cost

        #Conduct safer search for artist, considering the fact there might be ( )
        if "(" in product_vendor:

            #Obtain all chars up to the " (""
            product_vendor = product_vendor[:product_vendor.index("(") - 1]

        #First check to see if our vendor is an *actual* artist
        if product_vendor.title() in artist_info:
            
            product_vendor = product_vendor.title()

        else:
            

            #If our quick search of artist is not found, we need to look up
            #the artist aliases. LOOK UP EXACT MATCH (No titles)
            search_vendor = artist_lookup(product_vendor, artist_info)

            if search_vendor is not None:
                
                product_vendor = search_vendor

            #Our last resort search will be for Yorunmachi vended items,
            #If we cannot find the vendor at this point, we can assume we 
            #need to enter a new entry or make a fix
            else:

                search_vendor_item = item_lookup(product_name, item_info)

                if search_vendor_item:

                    product_vendor = search_vendor_item
                
                else:

                    print("ARTIST NOT FOUND, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                    continue

        #
        #Get associated information for the artist's product - handle differently dependent on the distribution type
        #

        if distribution_type == "Original":

            if product_vendor not in yne_original_dict:

                yne_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]

            else:

                yne_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])

        elif distribution_type == "Commercial":

            if product_vendor not in yne_collab_dict:

                yne_collab_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]

            else:

                yne_collab_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])


        elif distribution_type == "Collab":

            #We find the collaborator and add them to the collab dict
            collab_name = name_extract(product_name, artist_info)

            if not(collab_name):

                print("COLLABORATOR NOT FOUND, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                continue

            if collab_name not in yne_collab_dict:

                yne_collab_dict[collab_name] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]
            
            else:
                    
                yne_collab_dict[collab_name].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])

            #Handle if YNE was a vendor - we have to take collaborator's name out of list as it was already accounted for
            if type(product_vendor) == list:

                try:
                    product_vendor.remove(collab_name)
                except:

                    print("FIX NAME:", collab_name, product_vendor, product_name)

                #Recreate the one element list to a string
                product_vendor = product_vendor[0]

            #Next, we must find the product vendor as an original source, so they go in the original source dict
            if product_vendor not in yne_original_dict:

                yne_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit]]

            else:

                yne_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit])             


    #
    #Start Creating the Output Spreadsheet
    #
    ynm_artist_payments_dict = {}
    yne_artist_payments_dict = {}

    #Start with YNM
    ynm_original_dict = dict(sorted(ynm_original_dict.items()))

    artist_col = []
    item_col = []
    dist_type_col = []
    sales_col = []
    process_fee_col = []
    cost_col = []
    profit_col = []

    for key, val_list in ynm_original_dict.items():

        for val in val_list:
            artist_col.append(key)
            item_col.append(val[0])
            dist_type_col.append(val[1])
            sales_col.append(float(val[2]))
            process_fee_col.append(round(float(val[3]), 2))
            cost_col.append(round(float(val[4]),2))
            profit_col.append(float(val[5]))

    
    original_source_df = pd.DataFrame()
    original_source_df["Artist"] = artist_col
    original_source_df["Item"] = item_col
    original_source_df["Distribution Type"] = dist_type_col
    original_source_df["Total Sales"] = sales_col
    original_source_df["Processing Fee"] = process_fee_col
    original_source_df["Total Cost"] = cost_col
    original_source_df["Gross Profit"] = profit_col
    # original_source_df["Gross Profit"] = original_source_df["Total Sales"] - original_source_df["Total Cost"]
    original_source_df["SPE 40%"] = original_source_df["Gross Profit"] * 0.4
    original_source_df["SPE 50%"] = original_source_df["Gross Profit"] * 0.5
    original_source_df["SPE 60%"] = original_source_df["Gross Profit"] * 0.6




    ynm_collab_dict = dict(sorted(ynm_collab_dict.items()))

    artist_col = []
    item_col = []
    dist_type_col = []
    sales_col = []
    process_fee_col = []
    cost_col = []
    profit_col = []

    for key, val_list in ynm_collab_dict.items():
            
            for val in val_list:
                artist_col.append(key)
                item_col.append(val[0])
                dist_type_col.append(val[1])
                sales_col.append(float(val[2]))
                process_fee_col.append(round(float(val[3]), 2))
                cost_col.append(round(float(val[4]),2))
                profit_col.append(float(val[5]))

    collab_commercial_df = pd.DataFrame()
    collab_commercial_df["Artist"] = artist_col
    collab_commercial_df["Item"] = item_col
    collab_commercial_df["Distribution Type"] = dist_type_col
    collab_commercial_df["Total Sales"] = sales_col
    collab_commercial_df["Processing Fee"] = process_fee_col
    collab_commercial_df["Total Cost"] = cost_col
    collab_commercial_df["Gross Profit"] = profit_col
    # collab_commercial_df["Gross Profit"] = collab_commercial_df["Total Sales"] - collab_commercial_df["Total Cost"]
    collab_commercial_df["SPE 40%"] = collab_commercial_df["Gross Profit"] * 0.4
    collab_commercial_df["SPE 50%"] = collab_commercial_df["Gross Profit"] * 0.5
    collab_commercial_df["SPE 60%"] = collab_commercial_df["Gross Profit"] * 0.6


    #Next we go over YNE, doing the same thing
    yne_original_dict = dict(sorted(yne_original_dict.items()))

    artist_col = []
    item_col = []
    dist_type_col = []
    sales_col = []
    process_fee_col = []
    cost_col = []
    profit_col = []

    for key, val_list in yne_original_dict.items():

        for val in val_list:
            artist_col.append(key)
            item_col.append(val[0])
            dist_type_col.append(val[1])
            sales_col.append(float(val[2]))
            process_fee_col.append(round(float(val[3]), 2))
            cost_col.append(round(float(val[4]),2))
            profit_col.append(float(val[5]))

    
    yne_original_source_df = pd.DataFrame()
    yne_original_source_df["Artist"] = artist_col
    yne_original_source_df["Item"] = item_col
    yne_original_source_df["Distribution Type"] = dist_type_col
    yne_original_source_df["Total Sales"] = sales_col
    yne_original_source_df["Processing Fee"] = process_fee_col
    yne_original_source_df["Total Cost"] = cost_col
    yne_original_source_df["Gross Profit"] = profit_col
    # yne_original_source_df["Gross Profit"] = yne_original_source_df["Total Sales"] - yne_original_source_df["Total Cost"]
    yne_original_source_df["SPE 40%"] = yne_original_source_df["Gross Profit"] * 0.4
    yne_original_source_df["SPE 50%"] = yne_original_source_df["Gross Profit"] * 0.5
    yne_original_source_df["SPE 60%"] = yne_original_source_df["Gross Profit"] * 0.6


    yne_collab_dict = dict(sorted(yne_collab_dict.items()))

    artist_col = []
    item_col = []
    dist_type_col = []
    sales_col = []
    process_fee_col = []
    cost_col = []
    profit_col = []

    for key, val_list in yne_collab_dict.items():

        for val in val_list:
            artist_col.append(key)
            item_col.append(val[0])
            dist_type_col.append(val[1])
            sales_col.append(float(val[2]))
            process_fee_col.append(round(float(val[3]), 2))
            cost_col.append(round(float(val[4]),2))
            profit_col.append(float(val[5]))

    
    yne_collab_commercial_df = pd.DataFrame()
    yne_collab_commercial_df["Artist"] = artist_col
    yne_collab_commercial_df["Item"] = item_col
    yne_collab_commercial_df["Distribution Type"] = dist_type_col
    yne_collab_commercial_df["Total Sales"] = sales_col
    yne_collab_commercial_df["Processing Fee"] = process_fee_col
    yne_collab_commercial_df["Total Cost"] = cost_col
    yne_collab_commercial_df["Gross Profit"] = profit_col
    # yne_collab_commercial_df["Gross Profit"] = yne_collab_commercial_df["Total Sales"] - yne_collab_commercial_df["Total Cost"]
    yne_collab_commercial_df["SPE 40%"] = yne_collab_commercial_df["Gross Profit"] * 0.4
    yne_collab_commercial_df["SPE 50%"] = yne_collab_commercial_df["Gross Profit"] * 0.5
    yne_collab_commercial_df["SPE 60%"] = yne_collab_commercial_df["Gross Profit"] * 0.6





    with pd.ExcelWriter("Data/PayoutPrototype/YNM_Payout_Prototype.xlsx", engine="openpyxl") as writer:
        original_source_df.to_excel(writer, sheet_name="YNM Artist Payouts", index=False)
            
        # Workbook
        workbook = writer.book
        # Worksheet
        worksheet = workbook["YNM Artist Payouts"]

        # Create Styles to be used in the workbook
        bold_underline = Font(bold=True, underline="single")
        center = Alignment(horizontal="center")
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        new_header = ["Artist Name", "Item", "Distribution Type", "Total Value", "Processing Fee", "Cost", "Profit", "SPE 40%", "SPE 50%", "SPE 60%"]
        green_fill = PatternFill(start_color="00dc00", end_color="00dc00", fill_type="solid")

        # Apply the fill to cells where the artist changes
        i = 2  # start from 2 because 1 is the header
        total_profit = 0  # initialize total profit
        while i <= worksheet.max_row:
            if worksheet.cell(row=i, column=1).value != worksheet.cell(row=i-1, column=1).value:
                # Insert total profit for previous artist
                if i > 2:  # skip for the first artist
                    worksheet.insert_rows(i)
                    worksheet.cell(row=i, column=9).value = "Total Payout"
                    worksheet.cell(row=i, column=9).font = bold_underline
                    worksheet.cell(row=i, column=10).value = total_profit
                    worksheet.cell(row=i, column=10).font = bold_underline
                    i += 1  # skip the newly inserted row
                # Reset total profit for the new artist
                total_profit = 0
                new_artist = worksheet.cell(row=i, column=1).value
                worksheet.insert_rows(i)
                for j in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=i, column=j).fill = fill
                i += 1  # skip the newly inserted row
                worksheet.insert_rows(i)
                worksheet.cell(row=i, column=1).value = new_artist
                worksheet.cell(row=i, column=1).font = bold_underline  # make the artist name bold and underlined
                worksheet.cell(row=i, column=1).alignment = center
                # i += 1  # skip the newly inserted row
                # worksheet.insert_rows(i)
                for j, header in enumerate(new_header, start=1):
                    cell = worksheet.cell(row=i, column=j)
                    cell.value = header
                    cell.font = bold_underline  # make the header bold and underlined
                    cell.alignment = center
                i += 1  # skip the newly inserted row

                new_profit = 0
                if worksheet.cell(row=i, column=1).value not in ynm_artist_payments_dict.keys():

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit

            else:
                if worksheet.cell(row=i, column=1).value not in ynm_artist_payments_dict.keys():
                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit
            
            i += 1

        # Insert total profit for the last artist
        worksheet.insert_rows(i)
        worksheet.cell(row=i, column=9).value = "Total Payout"
        worksheet.cell(row=i, column=9).font = bold_underline
        worksheet.cell(row=i, column=10).value = total_profit
        worksheet.cell(row=i, column=10).font = bold_underline

        # Remove header
        worksheet.delete_rows(1)
    
        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


        collab_commercial_df.to_excel(writer, sheet_name="YNM Collab Payouts", index=False)
            
        # Workbook
        workbook = writer.book
        # Worksheet
        worksheet = workbook["YNM Collab Payouts"]


        # Apply the fill to cells where the artist changes
        i = 2  # start from 2 because 1 is the header
        total_profit = 0  # initialize total profit
        while i <= worksheet.max_row:
            if worksheet.cell(row=i, column=1).value != worksheet.cell(row=i-1, column=1).value:
                # Insert total profit for previous artist
                if i > 2:  # skip for the first artist
                    worksheet.insert_rows(i)
                    worksheet.cell(row=i, column=9).value = "Total Payout"
                    worksheet.cell(row=i, column=9).font = bold_underline
                    worksheet.cell(row=i, column=10).value = total_profit
                    worksheet.cell(row=i, column=10).font = bold_underline
                    i += 1  # skip the newly inserted row
                # Reset total profit for the new artist
                total_profit = 0
                new_artist = worksheet.cell(row=i, column=1).value
                worksheet.insert_rows(i)
                for j in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=i, column=j).fill = fill
                i += 1  # skip the newly inserted row
                worksheet.insert_rows(i)
                worksheet.cell(row=i, column=1).value = new_artist
                worksheet.cell(row=i, column=1).font = bold_underline  # make the artist name bold and underlined
                worksheet.cell(row=i, column=1).alignment = center
                # i += 1  # skip the newly inserted row
                # worksheet.insert_rows(i)
                for j, header in enumerate(new_header, start=1):
                    cell = worksheet.cell(row=i, column=j)
                    cell.value = header
                    cell.font = bold_underline  # make the header bold and underlined
                    cell.alignment = center
                i += 1  # skip the newly inserted row

                new_profit = 0
                if worksheet.cell(row=i, column=1).value not in ynm_artist_payments_dict.keys():

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit

            else:
                if worksheet.cell(row=i, column=1).value not in ynm_artist_payments_dict.keys():
                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    ynm_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit
            i += 1

        # Insert total profit for the last artist
        worksheet.insert_rows(i)
        worksheet.cell(row=i, column=9).value = "Total Payout"
        worksheet.cell(row=i, column=9).font = bold_underline
        worksheet.cell(row=i, column=10).value = total_profit
        worksheet.cell(row=i, column=10).font = bold_underline

        # Remove header
        worksheet.delete_rows(1)
    
        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        #Look over and see if we have any rollovers that we need to apply
        for rollover in rollover_info:

            print(rollover)

            #Are they an easy find in the dictionary?
            if rollover[1].title() in ynm_artist_payments_dict.keys():

                ynm_artist_payments_dict[rollover[1].title()] += round(float(rollover[2]),2)

            #Make sure that even if not found, we aren't under an alias
            else:

                search_vendor = artist_lookup(rollover[1], artist_info)

                if search_vendor is not None and search_vendor in ynm_artist_payments_dict.keys():

                    ynm_artist_payments_dict[search_vendor] += round(float(rollover[2]),2)

                else:

                    print("ARTIST NOT FOUND ALREADY, ADDING NEW ENTRY", rollover[1].title(), "==>", rollover[2])
                    ynm_artist_payments_dict[rollover[1].title()] = round(float(rollover[2]),2)

        print(ynm_artist_payments_dict)
        ynm_artist_payments_dict = dict(sorted(ynm_artist_payments_dict.items()))

        payments_df = pd.DataFrame()
        artist_col = []
        payment_col = []

        for key, val in ynm_artist_payments_dict.items():

            artist_col.append(key)
            payment_col.append(round(val, 2))

        payments_df["Artist"] = artist_col
        payments_df["Payment Due"] = payment_col

        #
        # Now we move onto YNE
        #


        yne_original_source_df.to_excel(writer, sheet_name="YNE Artist Payouts", index=False)
            
        # Workbook
        workbook = writer.book
        # Worksheet
        worksheet = workbook["YNE Artist Payouts"]

        # Create Styles to be used in the workbook
        # Define the fill colors
        red_fill = PatternFill(start_color="FF474C", end_color="FF474C", fill_type="solid")
        blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        bold_underline = Font(bold=True, underline="single")
        center = Alignment(horizontal="center")
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        new_header = ["Artist Name", "Item", "Distribution Type", "Total Value", "Processing Fee", "Cost", "Profit", "SPE 40%", "SPE 50%", "SPE 60%"]
        green_fill = PatternFill(start_color="00dc00", end_color="00dc00", fill_type="solid")

        # Apply the fill to cells where the artist changes
        i = 2  # start from 2 because 1 is the header
        total_profit = 0  # initialize total profit
        while i <= worksheet.max_row:
            if worksheet.cell(row=i, column=1).value != worksheet.cell(row=i-1, column=1).value:
                # Insert total profit for previous artist
                if i > 2:  # skip for the first artist
                    worksheet.insert_rows(i)
                    worksheet.cell(row=i, column=9).value = "Total Payout"
                    worksheet.cell(row=i, column=9).font = bold_underline
                    worksheet.cell(row=i, column=10).value = total_profit
                    worksheet.cell(row=i, column=10).font = bold_underline
                    i += 1  # skip the newly inserted row
                # Reset total profit for the new artist
                total_profit = 0
                new_artist = worksheet.cell(row=i, column=1).value
                worksheet.insert_rows(i)
                for j in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=i, column=j).fill = fill
                i += 1  # skip the newly inserted row
                worksheet.insert_rows(i)
                worksheet.cell(row=i, column=1).value = new_artist
                worksheet.cell(row=i, column=1).font = bold_underline  # make the artist name bold and underlined
                worksheet.cell(row=i, column=1).alignment = center
                # i += 1  # skip the newly inserted row
                # worksheet.insert_rows(i)
                for j, header in enumerate(new_header, start=1):
                    cell = worksheet.cell(row=i, column=j)
                    cell.value = header
                    cell.font = bold_underline  # make the header bold and underlined
                    cell.alignment = center
                i += 1  # skip the newly inserted row

                new_profit = 0
                if worksheet.cell(row=i, column=1).value not in yne_artist_payments_dict.keys():

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":
                        
                        try:
                            total_profit += worksheet.cell(row=i, column=10).value
                            worksheet.cell(row=i, column=10).fill = green_fill

                            new_profit = worksheet.cell(row=i, column=10).value

                        except:

                            print(worksheet.cell(row=i, column=1).value, "is not giving us a valid number with their item", worksheet.cell(row=i, column=2).value)
                            


                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit

            else:
                if worksheet.cell(row=i, column=1).value not in yne_artist_payments_dict.keys():
                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit
            
            i += 1

        # Insert total profit for the last artist
        worksheet.insert_rows(i)
        worksheet.cell(row=i, column=9).value = "Total Payout"
        worksheet.cell(row=i, column=9).font = bold_underline
        worksheet.cell(row=i, column=10).value = total_profit
        worksheet.cell(row=i, column=10).font = bold_underline

        # Remove header
        worksheet.delete_rows(1)
    
        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


        yne_collab_commercial_df.to_excel(writer, sheet_name="YNE Collab Payouts", index=False)
            
        # Workbook
        workbook = writer.book
        # Worksheet
        worksheet = workbook["YNE Collab Payouts"]


        # Apply the fill to cells where the artist changes

        i = 2  # start from 2 because 1 is the header
        total_profit = 0  # initialize total profit
        while i <= worksheet.max_row:
            if worksheet.cell(row=i, column=1).value != worksheet.cell(row=i-1, column=1).value:
                # Insert total profit for previous artist
                if i > 2:  # skip for the first artist
                    worksheet.insert_rows(i)
                    worksheet.cell(row=i, column=9).value = "Total Payout"
                    worksheet.cell(row=i, column=9).font = bold_underline
                    worksheet.cell(row=i, column=10).value = total_profit
                    worksheet.cell(row=i, column=10).font = bold_underline
                    i += 1  # skip the newly inserted row
                # Reset total profit for the new artist
                total_profit = 0
                new_artist = worksheet.cell(row=i, column=1).value
                worksheet.insert_rows(i)
                for j in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=i, column=j).fill = fill
                i += 1  # skip the newly inserted row
                worksheet.insert_rows(i)
                worksheet.cell(row=i, column=1).value = new_artist
                worksheet.cell(row=i, column=1).font = bold_underline  # make the artist name bold and underlined
                worksheet.cell(row=i, column=1).alignment = center
                # i += 1  # skip the newly inserted row
                # worksheet.insert_rows(i)
                for j, header in enumerate(new_header, start=1):
                    cell = worksheet.cell(row=i, column=j)
                    cell.value = header
                    cell.font = bold_underline  # make the header bold and underlined
                    cell.alignment = center
                i += 1  # skip the newly inserted row

                new_profit = 0
                if worksheet.cell(row=i, column=1).value not in yne_artist_payments_dict.keys():

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit

            else:
                if worksheet.cell(row=i, column=1).value not in yne_artist_payments_dict.keys():
                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] = new_profit

                else:

                    #Ensure that we still add the new profit to the total profit
                    if worksheet.cell(row=i, column=3).value == "Original":

                        total_profit += worksheet.cell(row=i, column=10).value
                        worksheet.cell(row=i, column=10).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=10).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Commercial":

                        total_profit += worksheet.cell(row=i, column=9).value
                        worksheet.cell(row=i, column=9).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=9).value
                    
                    elif worksheet.cell(row=i, column=3).value == "Collab":

                        total_profit += worksheet.cell(row=i, column=8).value
                        worksheet.cell(row=i, column=8).fill = green_fill

                        new_profit = worksheet.cell(row=i, column=8).value

                    yne_artist_payments_dict[worksheet.cell(row=i, column=1).value] += new_profit
            i += 1

        # Insert total profit for the last artist
        worksheet.insert_rows(i)
        worksheet.cell(row=i, column=9).value = "Total Payout"
        worksheet.cell(row=i, column=9).font = bold_underline
        worksheet.cell(row=i, column=10).value = total_profit
        worksheet.cell(row=i, column=10).font = bold_underline

        # Remove header
        worksheet.delete_rows(1)
    
        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        yne_artist_payments_dict = dict(sorted(yne_artist_payments_dict.items()))

        payments_df = pd.DataFrame()
        artist_col = []
        payment_col = []

        for key, val in yne_artist_payments_dict.items():

            artist_col.append(key)
            payment_col.append(round(val, 2))

        payments_df["Artist"] = artist_col
        payments_df["Payment Due"] = payment_col

        #Now we will make an aggregated sheet for all payouts both YNM and YNE

        #First we combine the dictionaries to get aggregated payments
        combined_artist_payments_dict = ynm_artist_payments_dict.copy()

        for key, val in yne_artist_payments_dict.items():

            if key in combined_artist_payments_dict.keys():

                combined_artist_payments_dict[key] += val

            else:

                combined_artist_payments_dict[key] = val

        
        combined_payments_df = pd.DataFrame()

        artist_col = []
        payment_col = []
        for key, val in combined_artist_payments_dict.items():
    
                artist_col.append(key)
                payment_col.append(round(val, 2))

        combined_payments_df["Artist"] = artist_col
        combined_payments_df["Payment Due"] = payment_col

        #Make new sheet for combined payments
        combined_payments_df.to_excel(writer, sheet_name="Combined Payouts", index=False)

        # Workbook
        workbook = writer.book
        # Worksheet
        worksheet = workbook["Combined Payouts"]

        # Add a new column "Paid?" on the left
        worksheet.insert_cols(1)
        worksheet.cell(row=1, column=1).value = "Paid?"
        worksheet.cell(row=1, column=1).font = Font(bold=True)

        # Add new columns "Payments Made", "Transaction ID", and "Notes" on the right
        for i, header in enumerate(["Payments Made", "Transaction ID", "Notes"], start=worksheet.max_column+1):
            worksheet.cell(row=1, column=i).value = header
            worksheet.cell(row=1, column=i).font = Font(bold=True)

        # Define the fill colors
        red_fill = PatternFill(start_color="FF474C", end_color="FF474C", fill_type="solid")

        #Apply to all except header
        worksheet.conditional_formatting.add(f'A2:A{worksheet.max_row}', CellIsRule(operator='equal', formula=['"y"'], fill=green_fill))
        worksheet.conditional_formatting.add(f'A2:A{worksheet.max_row}', CellIsRule(operator='equal', formula=['"r"'], fill=blue_fill))
        worksheet.conditional_formatting.add(f'A2:A{worksheet.max_row}', CellIsRule(operator='notEqual', formula=['"y"'], fill=red_fill))


        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


        #Now create a combined rollovers sheet
        combined_rollover_df = pd.DataFrame()

        reason_col = []
        artist_col = []
        rollover_col = []

        for key, val in combined_artist_payments_dict.items():

            if val < 0:

                reason_col.append("Credit")
                artist_col.append(key)
                rollover_col.append(val)

            elif val < 5:

                reason_col.append("< $5")
                artist_col.append(key)
                rollover_col.append(val)

        combined_rollover_df["Reason"] = reason_col
        combined_rollover_df["Artist"] = artist_col
        combined_rollover_df["Rollover Amount"] = rollover_col

        #Make new sheet for rollovers - Does not need to be styled except for fitting columns
        combined_rollover_df.to_excel(writer, sheet_name="Combined Next Month Rollovers", index=False)

        # Workbook
        workbook = writer.book

        # Worksheet
        worksheet = workbook["Combined Next Month Rollovers"]

        # AutoFit column width
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.0
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        #Create a pending rollovers sheet
        pending_rollover_df = pd.DataFrame()

        origin_col = []
        artist_col = []
        item_col = []
        dist_type_col = []
        sales_col = []
        process_fee_col = []
        cost_col = []
        profit_col = []
        ready_col = []

        for carryover in carry_pending_rollovers:

            origin_col.append(carryover[0])
            artist_col.append(carryover[1])
            item_col.append(carryover[2])
            dist_type_col.append(carryover[3])
            sales_col.append(float(carryover[4]))
            process_fee_col.append(float(carryover[5]))
            cost_col.append(float(carryover[6]))
            profit_col.append(float(carryover[7]))
            ready_col.append(carryover[8])

        pending_rollover_df["Origin"] = origin_col
        pending_rollover_df["Artist"] = artist_col
        pending_rollover_df["Item"] = item_col
        pending_rollover_df["Distribution Type"] = dist_type_col
        pending_rollover_df["Total Value"] = sales_col
        pending_rollover_df["Processing Fee"] = process_fee_col
        pending_rollover_df["Cost"] = cost_col
        pending_rollover_df["Profit"] = profit_col
        pending_rollover_df["Ready"] = ready_col

        pending_rollover_df.to_excel(writer, sheet_name="Pending Rollovers", index=False)

main()
