import sys
import os
curr_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(curr_dir)

import json
import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from ExcelRW.readcsv import read_csv
from datalookups import artist_lookup, identify_vendors, isPremium
from Models.PendingRollover import PendingRollover, parse_pending_rollovers

ynm_file_path = '/YNM/PayoutAutomator/Data/SheetPreprocessor/YNM_Sales_Final.csv'
ynm_financial_info = read_csv(ynm_file_path)

yne_file_path = '/YNM/PayoutAutomator/Data/SheetPreprocessor/YNE_Sales_Final.csv'
yne_financial_info = read_csv(yne_file_path)

convention_file_path = '/YNM/PayoutAutomator/Data/SheetPreprocessor/Convention_Sales_Final.csv'
convention_financial_info = read_csv(convention_file_path)

rollover_path = "/YNM/PayoutAutomator/Data/PayoutPrototype/Rollovers.csv"
rollover_info = read_csv(rollover_path)

pending_rollover_path = '/YNM/PayoutAutomator/Data/PayoutPrototype/Pending_Rollovers.csv'
pending_rollover_info = read_csv(pending_rollover_path)

ynm_financial_info, yne_financial_info, carry_pending_rollovers = parse_pending_rollovers(pending_rollover_info, ynm_financial_info, yne_financial_info)

with open('/YNM/PayoutAutomator/Data/artists.json') as fp:
    artist_info = json.load(fp)

def parse_product_information(store_financial_info: list):

    store_original_dict = {}
    store_collab_dict = {}

    #Iterate through each product that was sold this month for YNM
    for product in store_financial_info:
        product_vendors = product[1].strip("'").replace("''", "'")
        product_name = product[0].strip("'").replace("''", "'") #Clean up the string to be readable in json
        distribution_type = product[2]
        total_sales = float(product[10])
        processor_fee = float(product[11])
        total_cost = float(product[12])


        #Handle processing fee for negative sales, let us eat the processing fee for these, but not debit to the artist
        if total_sales < 0:
            processor_fee = 0

        gross_profit = total_sales - processor_fee - total_cost

        #Conduct safer search for artist, considering the fact there might be ( )
        if "(" in product_vendors and distribution_type != "Collab":

            #Obtain all chars up to the " (""
            product_vendor = product_vendors[:product_vendors.index("(") - 1]
        
        elif ")" in product_vendors and distribution_type == "Collab":

            try:
                #Obtain all chars up to the " (""
                product_vendor = product_vendors[:product_vendors.index("(") - 1]

                #Obtain all chars past the ") "
                second_product_vendor = product_vendors[product_vendors.index(") ") + 2:]

            except:

                print("ERROR IN COLLAB VENDOR NAME:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
        
        else:
            product_vendor = product_vendors

        #First check to see if our vendor is an *actual* artist
        if product_vendor.title() in artist_info:
            
            product_vendor = product_vendor.title()

        else:
            
            #If our quick search of artist is not found, we need to look up
            #the artist aliases. LOOK UP EXACT MATCH (No titles)
            search_vendor = artist_lookup(product_vendor, artist_info)

            if search_vendor is not None:
                
                product_vendor = search_vendor
            
            else:
                print("YNM ARTIST NOT FOUND, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
        
        #
        #Get associated information for the artist's product - handle differently dependent on the distribution type
        #

        if distribution_type == "Original":

            default_profit_split = 0.6

            if product_vendor not in store_original_dict:

                store_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split]]

            else:

                store_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit,default_profit_split ])
    
        elif distribution_type == "Digital":

            default_profit_split = 0.9

            if product_vendor not in store_original_dict:

                store_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split]]

            else:

                store_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split])

        elif distribution_type == "Charity":

            default_profit_split = 0

            if product_vendor not in store_collab_dict:

                store_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split]]

            else:

                store_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split])

        elif distribution_type == "Commercial" or distribution_type == "Book":

            default_profit_split = 0.5

            if product_vendor not in store_collab_dict:

                store_collab_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split]]

            else:

                store_collab_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split])


        elif distribution_type == "In-House":

            default_profit_split = 1

            if product_vendor not in store_original_dict:

                store_original_dict[product_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split]]

            else:

                store_original_dict[product_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split])

        elif distribution_type == "Collab":

            default_profit_split_collab = 0.2
            default_profit_split_artist = 0.4

            #First check to see if our vendor is an *actual* artist
            if second_product_vendor.title() in artist_info:
                
                second_product_vendor = second_product_vendor.title()
            else:

                #We find the collaborator and add them to the collab dict through a deeper search
                second_product_vendor = artist_lookup(second_product_vendor, artist_info)

            if not(second_product_vendor):

                print("COLLABORATOR NOT FOUND, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                continue


            #Perform a search to determine the roles of the first product vendor and second product venor
            if (product_vendor and second_product_vendor):

                vendor_identifications = identify_vendors(product_vendor, second_product_vendor, artist_info)

                if not(vendor_identifications):

                    print("COLLAB VENDORS COULD NOT BE IDENTIFIED, TAKE ACTION. SKIPPING FOR NOW:", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                    continue

                print("SUCCESS FOR COLLAB VENDORS:",vendor_identifications)

                artist_vendor = vendor_identifications["Artist"]
                collab_vendor = vendor_identifications["Collab"]

            else:

                print("One (or both) of the vendors were not found. Skipping for now.", product[1].title(), "==>",product_vendor, "(" + product_name + ")")
                continue

            #Determine if collaborator is a premium member
            if isPremium(collab_vendor, artist_info):

                default_profit_split_collab = 0.25

            if collab_vendor not in store_collab_dict:

                store_collab_dict[collab_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split_collab]]
            
            else:
                    
                store_collab_dict[collab_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split_collab])

            #Next, we must find the product vendor as an original source, so they go in the original source dict
            if artist_vendor not in store_original_dict:

                store_original_dict[artist_vendor] = [[product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split_artist]]

            else:

                store_original_dict[artist_vendor].append([product_name, distribution_type, total_sales, processor_fee, total_cost, gross_profit, default_profit_split_artist])  

    return store_original_dict, store_collab_dict

def construct_df(store_dict: dict):

    sorted_store_dict = dict(sorted(store_dict.items()))

    artist_col = []
    item_col = []
    dist_type_col = []
    sales_col = []
    process_fee_col = []
    cost_col = []
    profit_col = []
    profit_split_col = []
    client_profit = []

    for key, val_list in sorted_store_dict.items():

        for val in val_list:

            #skip if artist doesn't have any debit/credit for an item
            if val[2] == 0: 
                continue

            artist_col.append(key)
            item_col.append(val[0])
            dist_type_col.append(val[1])
            sales_col.append(float(val[2]))
            process_fee_col.append(round(float(val[3]), 2))
            cost_col.append(round(float(val[4]),2))
            profit_col.append(float(val[5]))
            profit_split_col.append(float(val[6]))
            client_profit.append(float(val[5]) * float(val[6]))

            #create logic to determine the profit cuts for each artist (dist_type_col and profit_col)

    store_source_df = pd.DataFrame()
    store_source_df["Artist"] = artist_col
    store_source_df["Item"] = item_col
    store_source_df["Distribution Type"] = dist_type_col
    store_source_df["Total Sales"] = sales_col
    store_source_df["Processing Fee"] = process_fee_col
    store_source_df["Total Cost"] = cost_col
    store_source_df["Gross Profit"] = profit_col
    store_source_df["SPE"] = profit_split_col
    store_source_df["Artist Profit"] = client_profit
    # store_source_df["SPE 40%"] = store_source_df["Gross Profit"] * 0.4
    # store_source_df["SPE 50%"] = store_source_df["Gross Profit"] * 0.5
    # store_source_df["SPE 60%"] = store_source_df["Gross Profit"] * 0.6
    # store_source_df["SPE 90%"] = store_source_df["Gross Profit"] * 0.9

    return pd.DataFrame(store_source_df)

def construct_payout_worksheet(store_source_df: pd.DataFrame, artists_payments_dict: dict, sheet_name: str):

    store_source_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
    # Workbook
    workbook = writer.book
    # Worksheet
    worksheet = workbook[sheet_name]

    # Create Styles to be used in the workbook
    bold_underline = Font(bold=True, underline="single")
    center = Alignment(horizontal="center")
    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    #TODO - change columns - consolidate columns
    new_header = ["Artist Name", "Item", "Distribution Type", "Total Value", "Processing Fee", "Cost", "Gross Profit", "SPE", "Client Profit"]

    # Apply the fill to cells where the artist changes
    i = 2  # start from 2 because 1 is the header
    total_profit = 0  # initialize total profit
    while i <= worksheet.max_row:
        if worksheet.cell(row=i, column=1).value != worksheet.cell(row=i-1, column=1).value:
            # Insert total profit for previous artist
            if i > 2:  # skip for the first artist
                worksheet.insert_rows(i)
                worksheet.cell(row=i, column=8).value = "Total Payout"
                worksheet.cell(row=i, column=8).font = bold_underline
                worksheet.cell(row=i, column=9).value = total_profit
                worksheet.cell(row=i, column=9).font = bold_underline
                i += 1  # skip the newly inserted row
            # Reset total profit for the new artist
            total_profit = 0
            new_artist = worksheet.cell(row=i, column=1).value
            worksheet.insert_rows(i)
            for j in range(1, worksheet.max_column + 1):
                worksheet.cell(row=i, column=j).fill = fill
            i += 1  # skip the newly inserted row
            worksheet.insert_rows(i)
            worksheet.cell(row=i, column=1).value = new_artist
            worksheet.cell(row=i, column=1).font = bold_underline  # make the artist name bold and underlined
            worksheet.cell(row=i, column=1).alignment = center
            for j, header in enumerate(new_header, start=1):
                cell = worksheet.cell(row=i, column=j)
                cell.value = header
                cell.font = bold_underline  # make the header bold and underlined
                cell.alignment = center
            i += 1  # skip the newly inserted row

        # Increment total_profit by the current row's value
        total_profit += worksheet.cell(row=i, column=9).value

        if worksheet.cell(row=i, column=1).value not in artists_payments_dict.keys():
            artists_payments_dict[worksheet.cell(row=i, column=1).value] = worksheet.cell(row=i, column=9).value
        else:
            artists_payments_dict[worksheet.cell(row=i, column=1).value] += worksheet.cell(row=i, column=9).value
        
        i += 1

    # Insert total profit for the last artist
    worksheet.insert_rows(i)
    worksheet.cell(row=i, column=8).value = "Total Payout"
    worksheet.cell(row=i, column=8).font = bold_underline
    worksheet.cell(row=i, column=9).value = total_profit
    worksheet.cell(row=i, column=9).font = bold_underline

    # Remove header
    worksheet.delete_rows(1)

    # AutoFit column width
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.0
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    return artists_payments_dict


def construct_convention_payout_worksheet(store_source_df: pd.DataFrame, artists_payments_dict: dict, sheet_name: str):

    pass

#Check to see if we updated non-automated pages
update_rollovers = input("Did you update the rollover sheet (y/n)? ")
update_pending_rollovers = input("Did you update the pending rollovers sheet (y/n)? ")
fix_payments_and_bugs = input("Did you resolve all payment fixes and bugs that were in the code? (y/n)? ")

if update_rollovers.lower() != "y" or update_pending_rollovers.lower() != "y" or fix_payments_and_bugs.lower() != "y":
    print("Please update the rollover and pending rollover sheets before running the script. Additionally ensure all bugs are fixed")
    sys.exit()

ynm_original_dict, ynm_collab_dict = parse_product_information(ynm_financial_info)
yne_original_dict, yne_collab_dict = parse_product_information(yne_financial_info)

ynm_original_df = construct_df(ynm_original_dict)
ynm_collab_df = construct_df(ynm_collab_dict)
yne_original_df = construct_df(yne_original_dict)
yne_collab_df = construct_df(yne_collab_dict)

#
#Start Creating the Output Spreadsheet
#
artists_payments_dict = {}

# Create Styles to be used in the workbook
bold_underline = Font(bold=True, underline="single")
center = Alignment(horizontal="center")
fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

#TODO - change columns 
new_header = ["Artist Name", "Item", "Distribution Type", "Total Value", "Processing Fee", "Cost", "Gross Profit", "SPE", "Client Profit"]
green_fill = PatternFill(start_color="00dc00", end_color="00dc00", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
purple_fill = PatternFill(start_color="800080", end_color="DB67DB", fill_type="solid")

with pd.ExcelWriter("Data/PayoutPrototype/YNM_Payout_Prototype.xlsx", engine="openpyxl") as writer:
    red_fill = PatternFill(start_color="FF474C", end_color="FF474C", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    artists_payments_dict = construct_payout_worksheet(ynm_original_df, artists_payments_dict, "YNM Artist Payouts")
    artists_payments_dict = construct_payout_worksheet(ynm_collab_df, artists_payments_dict, "YNM Collab Payouts")
    artists_payments_dict = construct_payout_worksheet(yne_original_df, artists_payments_dict, "YNE Artist Payouts")
    artists_payments_dict = construct_payout_worksheet(yne_collab_df, artists_payments_dict, "YNE Collab Payouts")
    aritsts_payment_dict = construct_convention_payout_worksheet(convention_financial_info, artists_payments_dict, "Convention Payouts")

  
    for rollover in rollover_info:

        rollover_artist = rollover[1].title()

        #Are they an easy find in the dictionary?
        if rollover_artist in artists_payments_dict.keys():

            artists_payments_dict[rollover_artist] += round(float(rollover[2]),2)

        #Make sure that even if not found, we aren't under an alias
        else:

            search_vendor = artist_lookup(rollover[1], artist_info)

            if search_vendor is not None and search_vendor in artists_payments_dict.keys():

                artists_payments_dict[search_vendor.title()] += round(float(rollover[2]),2)

            else:
                
                print("ARTIST NOT FOUND ALREADY, ADDING NEW ENTRY", rollover[1].title(), "==>", rollover[2])
                artists_payments_dict[rollover_artist] = round(float(rollover[2]),2)

    #Sort the dictionary by the artist's names
    artists_payments_dict = dict(sorted(artists_payments_dict.items()))

    combined_payments_df = pd.DataFrame()

    paid_status_col = []
    artist_col = []
    payment_col = []
    for key, val in artists_payments_dict.items():
            
            #skip if artist doesn't have any debit/credit
            if val == 0:
                continue
            
            if val < 5:
                paid_status_col.append("r")
            else: 
                paid_status_col.append("")

            artist_col.append(key)
            payment_col.append(round(val, 2))

    combined_payments_df["Paid?"] = paid_status_col
    combined_payments_df["Artist"] = artist_col
    combined_payments_df["Payment Due"] = payment_col

    #Make new sheet for combined payments
    combined_payments_df.to_excel(writer, sheet_name="Combined Payouts", index=False)

    # Workbook
    workbook = writer.book
    # Worksheet
    worksheet = workbook["Combined Payouts"]

    # # Add a new column "Paid?" on the left
    # worksheet.insert_cols(1)
    # worksheet.cell(row=1, column=1).value = "Paid?"
    # worksheet.cell(row=1, column=1).font = Font(bold=True)

    # Add new columns "Payments Made", "Transaction ID", and "Notes" on the right
    for i, header in enumerate(["Payments Made", "Transaction ID", "Notes"], start=worksheet.max_column+1):
        worksheet.cell(row=1, column=i).value = header
        worksheet.cell(row=1, column=i).font = Font(bold=True)

    # Define the fill colors
    red_fill = PatternFill(start_color="FF474C", end_color="FF474C", fill_type="solid")

    #Apply to all except header
    worksheet.conditional_formatting.add(f'A2:A{worksheet.max_row}', CellIsRule(operator='equal', formula=['"y"'], fill=green_fill))
    worksheet.conditional_formatting.add(f'A2:A{worksheet.max_row}', CellIsRule(operator='equal', formula=['"r"'], fill=blue_fill))
    worksheet.conditional_formatting.add(f'A2:A{worksheet.max_row}', CellIsRule(operator='notEqual', formula=['"y"'], fill=red_fill))

    # AutoFit column width
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.0
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    #Now create a combined rollovers sheet
    combined_rollover_df = pd.DataFrame()

    reason_col = []
    artist_col = []
    rollover_col = []

    for key, val in artists_payments_dict.items():

        #skip if artist doesn't have any debit/credit - zero balance for item
        if val == 0:
            continue

        if val < 0:

            reason_col.append("Credit")
            artist_col.append(key)
            rollover_col.append(val)

        elif val < 5:

            reason_col.append("< $5")
            artist_col.append(key)
            rollover_col.append(val)

    combined_rollover_df["Reason"] = reason_col
    combined_rollover_df["Artist"] = artist_col
    combined_rollover_df["Rollover Amount"] = rollover_col

    #Make new sheet for rollovers - Does not need to be styled except for fitting columns
    combined_rollover_df.to_excel(writer, sheet_name="Combined Next Month Rollovers", index=False)

    # Workbook
    workbook = writer.book

    # Worksheet
    worksheet = workbook["Combined Next Month Rollovers"]

    # AutoFit column width
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.0
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    #Create a pending rollovers sheet
    pending_rollover_df = pd.DataFrame()

    product_title_col = []
    product_vendor_col = []
    dist_type_col = []
    product_type_col = []
    net_quantity_col = []
    gross_sales_col = []
    discounts_col = []
    returns_col = []
    net_sales_col = []
    taxes_col = []
    total_sales_col = []
    total_cost_col = []
    processing_fee_col = []
    gross_profit_col = []
    origin_col = []
    ready_col = []

    for carryover in carry_pending_rollovers:

        product_title_col.append(carryover.product_title)
        product_vendor_col.append(carryover.product_vendor)
        dist_type_col.append(carryover.dist_type)
        product_type_col.append(carryover.product_type)
        net_quantity_col.append(carryover.net_quantity)
        gross_sales_col.append(carryover.gross_sales)
        discounts_col.append(carryover.discounts)
        returns_col.append(carryover.returns)
        net_sales_col.append(carryover.net_sales)
        taxes_col.append(carryover.taxes)
        total_sales_col.append(carryover.total_sales)
        total_cost_col.append(carryover.total_cost)
        processing_fee_col.append(carryover.processing_fee)
        gross_profit_col.append(carryover.gross_profit)
        origin_col.append(carryover.origin)
        ready_col.append(carryover.ready)

    pending_rollover_df["Product Title"] = product_title_col
    pending_rollover_df["Product Vendor"] = product_vendor_col
    pending_rollover_df["Distribution Type"] = dist_type_col
    pending_rollover_df["Product Type"] = product_type_col
    pending_rollover_df["Net Quantity"] = net_quantity_col
    pending_rollover_df["Gross Sales"] = gross_sales_col
    pending_rollover_df["Discounts"] = discounts_col
    pending_rollover_df["Returns"] = returns_col
    pending_rollover_df["Net Sales"] = net_sales_col
    pending_rollover_df["Taxes"] = taxes_col
    pending_rollover_df["Total Sales"] = total_sales_col
    pending_rollover_df["Total Cost"] = total_cost_col
    pending_rollover_df["Processing Fee"] = processing_fee_col
    pending_rollover_df["Gross Profit"] = gross_profit_col
    pending_rollover_df["Origin"] = origin_col
    pending_rollover_df["Ready"] = ready_col

    pending_rollover_df.to_excel(writer, sheet_name="Pending Rollovers", index=False)